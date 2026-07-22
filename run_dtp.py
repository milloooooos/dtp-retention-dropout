# -*- coding: utf-8 -*-
"""本地批量运行 dtp_engine，产出 Excel 分析报告（留存率/脱落率A/脱落率B/跨表原因）。"""
import pandas as pd, os
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import dtp_engine as E

OUT = 'output'
os.makedirs(OUT, exist_ok=True)
SRC = r'F:/厂家项目/阿斯利康/2026/半年复盘/英飞凡、利普卓、泰瑞沙、优赫得6.26.xlsx'
F25 = 'E:/下载内容/历史任务 - 2026-07-22T105636.443.xls'
F26 = 'E:/下载内容/历史任务 - 2026-07-22T105440.252.xls'

print('加载数据...')
# 项目药房 TOP 清单（可选）；若路径不存在则按销售底表角色列或默认项目处理
TOPF = r"D:\微信数据\xwechat_files\tonia852209598_370f\msg\file\2026-05\国控四川项目药房TOP清单-H1.xlsx"
tier = E.load_pharmacy_tier(TOPF) if os.path.exists(TOPF) else None
sales = E.load_sales(SRC, tier=tier)
followup = E.load_followup([F25, F26])
print(f'  销售: {len(sales)} 行, 患者 {sales["患者ID"].nunique()}, 时间 {sales["销售时间"].min().date()}~{sales["销售时间"].max().date()}')
print(f'  随访: {len(followup)} 行, 状态类分布 {followup["状态类"].value_counts().to_dict()}')

res = E.run_analysis(sales, followup, max_k=12, mult=3, with_patient_crossref=True, preset='H1_2026')

xlsx = os.path.join(OUT, 'DTP_留存率与脱落分析_v1.xlsx')
with pd.ExcelWriter(xlsx, engine='openpyxl') as xw:
    # 说明
    notes = pd.DataFrame({'说明': [
        'DTP 留存率 / 脱落率(A滚动+B累计沉默) / 跨表脱落原因 自动分析',
        '数据源：销售底表 + 随访历史任务表（2025H1+2026H1）',
        '',
        '【留存率·两套口径并存，按首购月分群=新患队列】',
        '  口径1(仅看购药时间)：Mk留存% = 该群首购后第k月"当月有购药"的患者占比。',
        '    特点：因多盒购买可非单调递减(囤3盒者次月不买第3月才回)；近期cohort窗口不足显示为空(右删失)。',
        '  口径2(结合说明书盒数覆盖)：每次购药覆盖=销售数量×每盒天数(说明书) 天，覆盖区间与第k月有交集即算留存。',
        '    特点：把囤药者的"在治"状态计入，曲线更平滑、更单调，更贴近真实在治率。两套对比看差值=囤药/断续行为。',
        '',
        '【统一时间窗】本报告所有指标(留存/脱落率B/脱落原因/DOT/新患同比/医生医院药房维度)共用一个时间窗，',
        '  由 run_analysis 的 preset 决定（本脚本默认 H1_2026 = 2026-01-01~2026-06-30；另可选 roll1y 回滚1年 / full 全量累计）。',
        '  脱落率B的判定终点=该窗终点(如 H1 为 2026-06-30)；脱落原因只统计 _dt 落在该窗内的随访记录。详见「0_窗口信息」sheet。',
        '',
        '【脱落率A·滚动】基准月M-2有购药、观察窗M-1∪M无购药→脱落。月度口径，窗内取月均。',
        '',
        '【脱落率B·累计沉默】末次购药距窗口终点 > 3×用药间隔→真停药。新近患者(首购在终点前3×间隔内)',
        '  免除右删失，故另给"已观察%"列（仅统计有充分观察期的患者）。',
        '',
        '【跨表关联】销售算出的脱落患者 → 随访表的脱落原因(细分类框架 + 一级分类)。',
        '  两表无共同患者主键，故：① 品牌层面关联为主(可靠)；② 患者级用复合键「姓名||药房全称」匹配',
        '  (同人不同药房视为不同人，防重名串号)；同名歧义仍存，置信度有限，仅供参考。',
        '  覆盖率：分母=销售脱落B患者；分子=随访表匹配上；其中区分「有记载原因」与「有记录但原因未填」。',
        '  低匹配率(<30%)报警：两张表很可能不是同一批患者(药品/项目/时间段不同)。',
        '  空原因≠无脱落：空原因单独计数，不混入任何细类。',
        '  随访表不含优赫得(0行)，故优赫得跨表原因为空(覆盖缺口，非错误)。',
        '',
        '【脱落原因分类·修正后】采用 dtp-churn-analyzer 的 churn_logic 框架：',
        '  10 个细类 + 一级(医生相关/患者相关/其他)，四原则(精确优先→关键词兜底→自上而下分区互斥→统一兜底)。',
        '  关键修正：医嘱相关(如"遵医嘱停药")优先于"自主停药/认知不足"判定，避免医生原因漏给"患者自主停药"。',
        '  细分类规则集中维护在 churn_logic.py，换项目只改该文件常量即可，统计逻辑不变。',
        '  归因维度仅用「细分类 + 一级分类」，不引入可控/不可控标签。',
    ]})
    notes.to_excel(xw, sheet_name='说明', index=False)
    if 'window_info' in res:
        res['window_info'].to_excel(xw, sheet_name='0_窗口信息', index=False)
    res['retention_overall'].to_excel(xw, sheet_name='1_留存率口径1_仅购药_整体', index=False)
    res['retention_by_brand'].to_excel(xw, sheet_name='2_留存率口径1_仅购药_分品种', index=False)
    res['retention_cov_overall'].to_excel(xw, sheet_name='1b_留存率口径2_盒数覆盖_整体', index=False)
    res['retention_cov_by_brand'].to_excel(xw, sheet_name='2b_留存率口径2_盒数覆盖_分品种', index=False)
    res['retention_by_brand_pharmacy'].to_excel(xw, sheet_name='2c_留存率_分品种×药房', index=False)
    res['dropout_A']['整体_月度'].to_excel(xw, sheet_name='3_脱落率A_整体月度', index=False)
    if '分品种_月度' in res['dropout_A']:
        res['dropout_A']['分品种_月度'].to_excel(xw, sheet_name='4_脱落率A_分品种', index=False)
    if '分药房_月度' in res['dropout_A']:
        res['dropout_A']['分药房_月度'].to_excel(xw, sheet_name='5_脱落率A_分药房', index=False)
    res['dropout_B_by_brand'].to_excel(xw, sheet_name='6_脱落率B_分品种', index=False)
    res['dropout_B_established_by_brand'].to_excel(xw, sheet_name='6b_脱落率B_已观察', index=False)
    if 'crossref_brand' in res:
        res['crossref_brand'].to_excel(xw, sheet_name='7_跨表关联_品牌层面', index=False)
    if 'dropout_reason_by_pharmacy' in res:
        res['dropout_reason_by_pharmacy'].to_excel(xw, sheet_name='7b_脱落原因_药房×品种', index=False)
    for _k, _n in [('dropout_reason_detail', '7b1_脱落原因_细分类分布'),
                   ('dropout_reason_lvl1', '7b2_脱落原因_一级汇总'),
                   ('dropout_reason_by_brand', '7b2b_脱落原因_按品种拆分'),
                   ('dropout_reason_meta', '7b3_脱落原因_覆盖概览'),
                   ('crossref_coverage', '8a_跨表覆盖率概览'),
                   ('patient_level_detail', '8b_患者级双视角明细')]:
        _df = res.get(_k)
        if isinstance(_df, pd.DataFrame) and not _df.empty:
            _df.to_excel(xw, sheet_name=_n, index=False)
    if 'crossref_patient' in res:
        res['crossref_patient'].to_excel(xw, sheet_name='8_跨表关联_患者级近似', index=False)
    res['action_map'].to_excel(xw, sheet_name='9_脱落原因行动建议', index=False)
    # 新增：DOT / 新患 / 复购 / 维度分析
    res['dot_decomposition'].to_excel(xw, sheet_name='10_DOT分解_老患vs新患', index=False)
    res['new_patient_monthly'].to_excel(xw, sheet_name='11_新患月度趋势', index=False)
    res['new_patient_pharmacy_decline'].to_excel(xw, sheet_name='12_新患下降最大药房', index=False)
    res['old_patient_multi_box'].to_excel(xw, sheet_name='13_老患多盒行为', index=False)
    res['repurchase_decomposition'].to_excel(xw, sheet_name='14_复购率分解', index=False)
    res['hospital_dimension'].to_excel(xw, sheet_name='15_医院维度', index=False)
    res['doctor_top1'].to_excel(xw, sheet_name='16_医生维度_最大患者量', index=False)
    res['doctor_top5'].to_excel(xw, sheet_name='17_医生维度_TOP5', index=False)
    res['doctor_low_dot_watch'].to_excel(xw, sheet_name='18_医生维度_低DOT重点', index=False)
    res['pharmacy_dimension'].to_excel(xw, sheet_name='19_项目药房_风险', index=False)
    res['drill_hospital_doctor'].to_excel(xw, sheet_name='20_钻取_异常医院拖后腿医生', index=False)
    res['communication_list'].to_excel(xw, sheet_name='21_重点关注清单', index=False)
    res['improvement_actions'].to_excel(xw, sheet_name='22_改进措施与责任', index=False)

# 简单样式
from openpyxl import load_workbook
wb = load_workbook(xlsx)
hdr_fill = PatternFill('solid', fgColor='4472C4')
for ws in wb.worksheets:
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 10), 40)
    ws.freeze_panes = 'A2'
wb.save(xlsx)

print('\n已生成:', xlsx)
print('\n=== 留存率口径1(仅购药时间) 整体(首3行) ===')
print(res['retention_overall'].head(3).to_string(index=False))
print('\n=== 留存率口径2(盒数覆盖) 整体(首3行) ===')
print(res['retention_cov_overall'].head(3).to_string(index=False))
print('\n映射表 unresolved(需人工补录):', E._ALIAS['unresolved'])
print('\n=== 脱落率B 分品种(全部) ===')
print(res['dropout_B_by_brand'].to_string(index=False))
print('\n=== 脱落率B 分品种(已观察) ===')
print(res['dropout_B_established_by_brand'].to_string(index=False))
if 'crossref_brand' in res:
    print('\n=== 跨表关联 品牌层面 ===')
    print(res['crossref_brand'].to_string(index=False))
